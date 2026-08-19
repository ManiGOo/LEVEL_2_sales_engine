"""Excel report generator for the sales team.

Reads company + signal data from the shared Pharma DB (scraped data in the
``sdr_data`` schema, produced by the Sentinel scraper) via ``sentinel_service``
and renders it into a styled two-sheet workbook:

- **General List**: one row per company — Company, Location, NSQ Alert, Rating,
  derived Status & Description, plus the full detailed scoring breakdown
  (paper-based bonuses, 2026-mandate, recency, repeat-offender, web evidence).
- **NSQ**: the monthly-extracted "Not of Standard Quality" list — one row per
  NSQ record with drug / batch / reason / reporting month and the same detailed
  scoring.

This module only converts data into an .xlsx; all data is read directly from the
shared database — no calls to the scraper.
"""

import io
import re
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services import sentinel_service

# ---------------------------------------------------------------------------
# Location extraction
# ---------------------------------------------------------------------------
# Address markers that appear right after a trading name inside the raw CDSCO
# manufacturer string (mirrors the scraper's company_names.ADDRESS_MARKERS,
# now available locally in app.scraper.names).
_ADDRESS_MARKERS = re.compile(
    r"(?i)"
    r"\bplot\s*(?:no\.?|no)?\s*:?|"
    r"\bkhasra\s*(?:no\.?)?|"
    r"\bsurvey\s*(?:no\.?)?|"
    r"\bsurvery\s*(?:no\.?)?|"
    r"\bk\.?\s*h\.?\s*no\.?|"
    r"\bs\.?\s*(?:f\.?\s*)?no\.?|"
    r"\bsector[- ]?\d|"
    r"\bphase[- ]?[iv\d]|"
    r"\bvill[e]?\b|"
    r"\bvillage\b|"
    r"\btehsil\b|"
    r"\btaluk\b|"
    r"\bdistt?\.?\b|"
    r"\bn[h]-?\s*\d|"
    r"\broad\b|"
    r"\brd\.?\b|"
    r"\bstreet\b|"
    r"\b\#\s?\d|"
    r"\b\d{6}\b|"
    r"\b(?:gidc|sidcul|hpsidc|hsiidc|sidco|ida|epip|sldc)\b|"
    r"\bindustrial\s*(?:area|estate|park|growth)\b|"
    r"\bindl\.?\s*area\b|"
    r"\bfactory\b|"
    r"\bat\s*[:.]"
)

_PAREN = re.compile(r"\([^)]*\)")


def _norm_text(raw: str) -> str:
    if not raw:
        return ""
    return re.sub(r"\s+", " ", raw.replace("\n", ", ")).strip(" ,;:.-")


def extract_location(mfr_raw: str, company_name: str = "") -> str:
    """Best-effort location from the raw CDSCO manufacturer string.

    The raw string embeds the plant address after the trading name, e.g.
    ``"Hanuchem Laboratories, Plot No 13, Sector-5, Industrial Area,
    Parwanoo, Distt. Solan (H.P)"`` -> ``"Plot No 13, Sector-5, Industrial
    Area, Parwanoo, Distt. Solan (H.P)"``.
    """
    raw = _norm_text(mfr_raw)
    if not raw:
        return ""

    # Prefer cutting right after the cleaned company name (exact anchor).
    name = _norm_text(company_name)
    if name:
        m = re.search(re.escape(name), raw, re.IGNORECASE)
        if m:
            tail = raw[m.end():].lstrip(" ,;:.-")
            if tail:
                return re.sub(r"\s+", " ", tail).strip(" ,;:.-")

    # Fallback: cut at the first address marker.
    m = _ADDRESS_MARKERS.search(raw)
    if m:
        tail = raw[m.start():].strip(" ,;:.-")
        if tail:
            return re.sub(r"\s+", " ", tail).strip(" ,;:.-")
    return ""


# ---------------------------------------------------------------------------
# Derived fields
# ---------------------------------------------------------------------------

def _nsq_alert(card: dict) -> str:
    """Whether the company has any NSQ (Not of Standard Quality) records."""
    count = sum(
        1
        for ev in [card] + list(card.get("events") or [])
        if (ev.get("event_type") or "") == "NSQ_DRUG"
    )
    return f"Yes ({count})" if count else "No"


def _clean_reason(card: dict) -> str:
    reason = (card.get("raw_details") or {}).get("reason", "") or ""
    return re.sub(r"\s+", " ", reason).strip()


def _reporting_month(ev: dict) -> str:
    """CDSCO reporting month when stored, else derived from the event date."""
    month = (ev.get("raw_details") or {}).get("dt_reporting_month_year", "")
    if month:
        return str(month)
    event_date = ev.get("event_date") or ""
    if len(str(event_date)) >= 7:
        return str(event_date)[:7]
    return ""


def _derive_status(card: dict) -> str:
    """Status derived purely from the signal data."""
    statuses = []
    analysis = card.get("llm_analysis") or {}
    if (card.get("event_type") or "") == "SPURIOUS_DRUG":
        statuses.append("Spurious drug")
    paper_class = (card.get("paper_assessment") or {}).get("class", "")
    if paper_class == "explicit":
        statuses.append("Paper-QMS (explicit)")
    elif paper_class == "deductive":
        statuses.append("Paper-QMS (deductive)")
    if any(analysis.get(k) for k in ("violates_rule_96", "violates_sub_rule_7", "violates_schedule_h2")):
        statuses.append("2026 mandate")
    if (card.get("score_breakdown") or {}).get("prior_events"):
        statuses.append("Repeat offender")
    if not statuses:
        statuses.append("NSQ alert" if (card.get("event_type") or "") == "NSQ_DRUG" else "Monitor")
    return ", ".join(statuses)


# ---------------------------------------------------------------------------
# Data fetch
# ---------------------------------------------------------------------------

async def _fetch_all(**params) -> list:
    """Page through the Sentinel signals endpoint until exhausted."""
    items = []
    page = 1
    params.setdefault("page_size", 200)
    while page <= 50:
        data = await sentinel_service.get_signals(page=page, **params)
        batch = data.get("items") or []
        items.extend(batch)
        pages = data.get("pages") or 1
        if page >= pages or not batch:
            break
        page += 1
    return items


# ---------------------------------------------------------------------------
# Sheet helpers
# ---------------------------------------------------------------------------

_TITLE_FONT = Font(bold=True, size=16, color="1F2937")
_SUBTITLE_FONT = Font(size=10, color="6B7280")
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
_BODY_FONT = Font(size=10, color="111827")
_MONO_FONT = Font(size=10, color="111827")
_ZEBRA_FILL = PatternFill("solid", fgColor="F3F4F6")
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_NSQ_YES_FILL = PatternFill("solid", fgColor="FEE2E2")
_NSQ_NO_FILL = PatternFill("solid", fgColor="ECFDF5")
_SPURIOUS_FILL = PatternFill("solid", fgColor="FDE68A")
_PAPER_FILL = PatternFill("solid", fgColor="DBEAFE")
_SCORE_HI_FILL = PatternFill("solid", fgColor="D1FAE5")

_GENERAL_HEADERS = [
    "Company",
    "Location",
    "NSQ Alert",
    "Rating (Peak Score)",
    "Max Possible Score",
    "Status",
    "Description",
    "Base Score",
    "Paper Bonus",
    "Paper Class",
    "Mandate Bonus",
    "Recency Weight",
    "Repeat-Offender Bonus",
    "Web Evidence Bonus",
    "Prior Events",
    "Latest Report Date",
]

_NSQ_HEADERS = [
    "Company",
    "Location",
    "Drug",
    "Batch",
    "Description",
    "Reporting Month",
    "Event Type",
    "Reported By",
    "Rating (Score)",
    "Base Score",
    "Paper Bonus",
    "Paper Class",
    "Mandate Bonus",
    "Recency Weight",
    "Repeat-Offender Bonus",
    "Web Evidence Bonus",
    "Report Date",
]


def _wrap_widths(headers: list, rows: list, cap: int = 48) -> list:
    widths = []
    for j, header in enumerate(headers):
        longest = len(str(header))
        for row in rows:
            if j < len(row):
                longest = max(longest, len(str(row[j] or "")))
        widths.append(min(longest + 2, cap))
    return widths


def _build_sheet(wb: Workbook, name: str, title: str, subtitle: str, headers: list, rows: list):
    ws = wb.create_sheet(title=name)
    ncols = len(headers)
    last_col = get_column_letter(ncols)

    # Title + subtitle banner
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(1, 1, title)
    title_cell.font = _TITLE_FONT
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{last_col}2")
    sub_cell = ws.cell(2, 1, subtitle)
    sub_cell.font = _SUBTITLE_FONT
    ws.row_dimensions[2].height = 16

    hdr_row = 3
    for j, header in enumerate(headers, start=1):
        cell = ws.cell(hdr_row, j, header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[hdr_row].height = 30

    # Body rows with zebra striping
    for i, row in enumerate(rows):
        r = hdr_row + 1 + i
        for j, value in enumerate(row, start=1):
            cell = ws.cell(r, j, value)
            cell.font = _BODY_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(
                vertical="top",
                horizontal="center" if headers[j - 1].lower() not in ("company", "location", "description", "drug", "status") else "left",
                wrap_text=headers[j - 1].lower() in ("company", "location", "description", "status", "drug"),
            )
            if i % 2 == 1:
                cell.fill = _ZEBRA_FILL

    # Column widths + freeze + autofilter
    for j, w in enumerate(_wrap_widths(headers, rows), start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{hdr_row + 1}"
    if rows:
        ws.auto_filter.ref = f"A{hdr_row}:{last_col}{hdr_row + len(rows)}"
    ws.sheet_view.showGridLines = False
    return ws


def _highlight_flags(ws, headers: list, rows: list, hdr_row: int = 3):
    """Row-level highlight fills for key status columns."""
    col = {h: i + 1 for i, h in enumerate(headers)}
    body_start = hdr_row + 1
    for i, row in enumerate(rows):
        r = body_start + i
        # NSQ alert column
        if "NSQ Alert" in col:
            cell = ws.cell(r, col["NSQ Alert"])
            if str(row[col["NSQ Alert"] - 1]).startswith("Yes"):
                cell.fill = _NSQ_YES_FILL
            else:
                cell.fill = _NSQ_NO_FILL
        # Spurious / paper rows
        if "Status" in col:
            status = str(row[col["Status"] - 1])
            if "Spurious" in status:
                for j in range(1, len(headers) + 1):
                    ws.cell(r, j).fill = _SPURIOUS_FILL
            elif "Paper-QMS" in status:
                for j in range(1, len(headers) + 1):
                    ws.cell(r, j).fill = _PAPER_FILL
        # High rating
        if "Rating (Peak Score)" in col:
            score = row[col["Rating (Peak Score)"] - 1]
            if isinstance(score, (int, float)) and score >= 80:
                ws.cell(r, col["Rating (Peak Score)"]).fill = _SCORE_HI_FILL


# ---------------------------------------------------------------------------
# Row builders
# ---------------------------------------------------------------------------

def _general_rows(cards: list) -> list:
    rows = []
    for card in cards:
        bd = card.get("score_breakdown") or {}
        raw = card.get("raw_details") or {}
        rows.append([
            card.get("company_name") or "",
            extract_location(raw.get("manufacturer", ""), card.get("company_name") or ""),
            _nsq_alert(card),
            card.get("score"),
            card.get("max_possible_score"),
            _derive_status(card),
            _clean_reason(card),
            bd.get("base"),
            bd.get("paper_bonus"),
            bd.get("paper_bonus_class") or "",
            bd.get("mandate_bonus"),
            bd.get("recency_weight"),
            bd.get("repeat_offender_bonus"),
            bd.get("web_evidence_bonus"),
            bd.get("prior_events"),
            card.get("event_date") or "",
        ])
    return rows


def _nsq_rows(events: list) -> list:
    rows = []
    for ev in events:
        bd = ev.get("score_breakdown") or {}
        raw = ev.get("raw_details") or {}
        rows.append([
            ev.get("company_name") or "",
            extract_location(raw.get("manufacturer", ""), ev.get("company_name") or ""),
            raw.get("drug_name") or "",
            raw.get("batch_no") or "",
            _clean_reason(ev),
            _reporting_month(ev),
            ev.get("event_type") or "",
            ev.get("reported_by") or "",
            ev.get("score"),
            bd.get("base"),
            bd.get("paper_bonus"),
            bd.get("paper_bonus_class") or "",
            bd.get("mandate_bonus"),
            bd.get("recency_weight"),
            bd.get("repeat_offender_bonus"),
            bd.get("web_evidence_bonus"),
            ev.get("event_date") or "",
        ])
    return rows


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

async def build_report() -> bytes:
    """Fetch company + NSQ data from Sentinel and return a styled .xlsx blob."""
    companies = await _fetch_all(group_by="company", page_size=200)
    nsq_events = await _fetch_all(event_type="NSQ_DRUG", page_size=200)

    companies.sort(key=lambda c: (c.get("score") or 0), reverse=True)
    nsq_events.sort(key=lambda e: (e.get("score") or 0), reverse=True)

    now = datetime.now()
    stamp = now.strftime("%d %b %Y %H:%M")
    month = now.strftime("%B %Y")

    general_rows = _general_rows(companies)
    nsq_rows = _nsq_rows(nsq_events)

    wb = Workbook()
    wb.remove(wb.active)

    ws_general = _build_sheet(
        wb,
        "General List",
        "AIVOA Sentinel — Company Intelligence",
        f"All companies ranked by lead score · generated {stamp} · {len(general_rows)} companies",
        _GENERAL_HEADERS,
        general_rows,
    )
    _highlight_flags(ws_general, _GENERAL_HEADERS, general_rows)

    ws_nsq = _build_sheet(
        wb,
        "NSQ",
        f"NSQ List — {month}",
        f"Monthly extraction of CDSCO Not-of-Standard-Quality records · generated {stamp} · {len(nsq_rows)} records",
        _NSQ_HEADERS,
        nsq_rows,
    )
    _highlight_flags(ws_nsq, _NSQ_HEADERS, nsq_rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
